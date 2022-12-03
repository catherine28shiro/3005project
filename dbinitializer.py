import psycopg2
from faker import Faker
from random import randint, randrange
import openpyxl
from datetime import datetime
from math import ceil

def fakeEmail(p):
    return p[:10].replace(' ','_') + '@gmail.com'
def fakeBank():
    return str(randint(4000,4999))+str(randint(1000,9999))+str(randint(1000,9999))+str(randint(1000,9999))
def expDateFaker():
    date = Faker().date_between_dates(date_start=datetime(2023,1,1), date_end=datetime(2025,12,31))
    return date.strftime('%m/%y')

class Book():
    def __init__(self,title,author,price,publisher,page_count,genres,ISBN,storage):
        self.title = title
        self.authors = author
        self.price = price
        self.publisher = publisher
        self.page_count = page_count
        self.genres = genres
        self.ISBN = ISBN
        self.storage = storage
    def __str__(self):
        return f"{self.title} by {self.authors}"

# connect to the database
hostname = 'peanut.db.elephantsql.com'
database = 'iogkcxzz'
username='iogkcxzz'
pwd='4BgZjz6v_2N1448Ftv1TQGLR5VKKT1ab'

conn = None
cur = None

# import data
def readFile():
    # get info from xlsx file
    book_set = openpyxl.load_workbook("book_dataset.xlsx")
    sh = book_set.active
    publisher_set = set()
    genre_set = set()
    author_set = set()
    bookshelf = []
    for i in range(2, sh.max_row+1):
        title = sh.cell(row=i, column=1).value.strip()
        authors = sh.cell(row=i, column=2).value.strip().split(',')
        price = sh.cell(row=i, column=3).value
        publisher = sh.cell(row=i, column=4).value.strip()
        page_count = sh.cell(row=i, column=5).value
        genres = sh.cell(row=i, column=6).value.strip().split(',')
        ISBN = sh.cell(row=i, column=7).value
        storage = randint(40, 50)
        book = Book(title,authors,price,publisher,page_count,genres,ISBN,storage)
        bookshelf.append(book)
        
        # check for publisher
        publisher_set.add(publisher)
        # check for genre
        for g in genres:
            genre_set.add(g.strip())
        for a in authors:
            author_set.add(a.strip())
    return publisher_set,genre_set,author_set,bookshelf

#get info from file
publisher_set,genre_set,author_set,bookshelf=readFile()

def ddl():
    print('ddl')
    #DDL
    create_utility_table='''create table IF NOT EXISTS UTILITY
        (MONTH          varchar(2)      NOT NULL UNIQUE,
        COST            numeric(8,2)    NOT NULL,        
        primary key (MONTH)
    );'''
 
    create_publisher_table='''create table IF NOT EXISTS PUBLISHER
        (NAME           varchar(60)     NOT NULL UNIQUE,
        ADDRESS         varchar(100),
        EMAIL           varchar(100),
        PHONE           varchar(20),
        BANK_ACC        varchar(50),
        PID             int             NOT NULL UNIQUE,
        primary key (PID)
    );'''
 
    create_book_table = '''create table IF NOT EXISTS BOOK
        (TITLE		            varchar(300)	NOT NULL,		
        PRICE		            numeric(7,2)	NOT NULL,
        PAGE_COUNT	            int             NOT NULL,
        ISBN	                varchar(13)     UNIQUE NOT NULL,
        STORAGE		            int             NOT NULL,
        BID                     int             UNIQUE NOT NULL,
        PID                     int,
        SALESPERCENT_TO_PUB     numeric(3,2)    NOT NULL,
        primary key (BID),
        foreign key (PID) references PUBLISHER (PID) ON DELETE CASCADE
	);'''
    create_genre_table = '''create table IF NOT EXISTS GENRE
        (genre_type     varchar(30) NOT NULL UNIQUE,
        GID             int         NOT NULL UNIQUE,
        primary key (GID)
    );'''
    create_typeof_table='''create table IF NOT EXISTS TYPEOF
        (GID    int     NOT NULL,
        BID     int     NOT NULL,
        primary key (BID,GID),
        foreign key (GID) references GENRE (GID) ON DELETE CASCADE,
        foreign key (BID) references BOOK (BID) ON DELETE CASCADE
    );'''
    create_author_table='''create table IF NOT EXISTS AUTHOR
        (AID    int         NOT NULL UNIQUE,
        NAME    varchar(50) NOT NULL,
        primary key (AID)
    );'''
    create_write_table='''create table IF NOT EXISTS WRITE
        (AID    int         NOT NULL,
        BID     int         NOT NULL,
        primary key (BID,AID),
        foreign key (AID) references AUTHOR (AID) ON DELETE CASCADE,
        foreign key (BID) references BOOK (BID) ON DELETE CASCADE
    );'''

    create_customer_table='''create table IF NOT EXISTS CUSTOMER
        (CID            int         NOT NULL UNIQUE,
        USERNAME        varchar(50) NOT NULL UNIQUE,
        PASSWORD        varchar(15) NOT NULL,  
        primary key (CID)  
    );'''
    create_cart_table='''create table IF NOT EXISTS SHOPPING_CART
        (CID            int         NOT NULL,  
        BID             int         NOT NULL,
        amount          int         NOT NULL,
        primary key (CID,BID),
        foreign key (BID) references BOOK (BID) ON DELETE CASCADE
    );'''
    create_card_table='''create table IF NOT EXISTS PAYMENT_CARD
        (CARD_NUM        varchar(16),
        CVV             varchar(3),
        EXP_DATE        varchar(5),
        BILL_NAME       varchar(50),
        BILL_ADDRESS    varchar(100),
        PAYMENT_TYPE    varchar(10),
        primary key (CARD_NUM)
    );'''
    create_order_table='''create table IF NOT EXISTS BOOK_ORDER
        (OID            int         NOT NULL,
        CARRIER         varchar(30),
        CID             int         NOT NULL, 
        SHIP_NAME       varchar(50),
        SHIP_ADDRESS    varchar(100),
        CARD_NUM        varchar(16),
        SHIPMENT_STATUS varchar(70),
        SDATE           DATE,
        primary key (OID),
        foreign key (CID) references CUSTOMER (CID) ON DELETE CASCADE,
        foreign key (CARD_NUM) references PAYMENT_CARD (CARD_NUM) ON DELETE CASCADE   
    );'''
    
    create_contain_table='''create table IF NOT EXISTS CONTAIN
        (OID            int         NOT NULL,
        BID             int         NOT NULL,
        AMOUNT          int         NOT NULL,
        primary key (OID,BID),
        foreign key (BID) references BOOK (BID) ON DELETE CASCADE,
        foreign key (OID) references BOOK_ORDER (OID) ON DELETE CASCADE
    );'''

    # the table that send email to publisher when book storage is less than 10
    create_email_pub_table='''create table IF NOT EXISTS ORDER_FROM_PUBLISHER
        (PUB_ORDER_ID   int             NOT NULL UNIQUE,
        ORDER_AMOUNT    int             NOT NULL,
        PID             int             NOT NULL,
        BID             int             NOT NULL,       
        DATE            DATE            NOT NULL,     
        primary key (PUB_ORDER_ID),
        foreign key (BID) references BOOK (BID) ON DELETE CASCADE,
        foreign key (PID) references PUBLISHER (PID) ON DELETE CASCADE
    );'''
  
    cur.execute('drop table IF EXISTS PUBLISHER CASCADE')
    cur.execute('drop table IF EXISTS BOOK CASCADE')
    cur.execute('drop table IF EXISTS AUTHOR CASCADE')
    cur.execute('drop table IF EXISTS GENRE CASCADE')
    cur.execute('drop table IF EXISTS TYPEOF CASCADE')
    cur.execute('drop table IF EXISTS WRITE CASCADE')
    cur.execute('drop table IF EXISTS BOOK_ORDER CASCADE')
    cur.execute('drop table IF EXISTS CUSTOMER CASCADE')
    cur.execute('drop table IF EXISTS CONTAIN CASCADE')
    cur.execute('drop table IF EXISTS SHOPPING_CART CASCADE')
    cur.execute('drop table IF EXISTS UTILITY CASCADE')
    cur.execute('drop table IF EXISTS PAYMENT_CARD CASCADE')
    cur.execute('drop table IF EXISTS ORDER_FROM_PUBLISHER CASCADE')

 
    cur.execute(create_publisher_table)
    cur.execute(create_book_table)
    cur.execute(create_genre_table)
    cur.execute(create_author_table)
    cur.execute(create_typeof_table)
    cur.execute(create_write_table)
    cur.execute(create_customer_table)
    cur.execute(create_card_table)
    cur.execute(create_order_table)
    cur.execute(create_contain_table)
    cur.execute(create_cart_table)
    cur.execute(create_utility_table)
    cur.execute(create_email_pub_table)
    
    conn.commit()

def createSequence():
    cur.execute('''CREATE SEQUENCE pidcreater
    AS BIGINT
    INCREMENT BY 1
    MINVALUE 100001
    NO MAXVALUE;
    ''')
    cur.execute('''CREATE SEQUENCE gidcreater
    AS BIGINT
    INCREMENT BY 1
    MINVALUE 200001
    NO MAXVALUE;
    ''')
    cur.execute('''CREATE SEQUENCE aidcreater
    AS BIGINT
    INCREMENT BY 1
    MINVALUE 300001
    NO MAXVALUE;
    ''')
    cur.execute('''CREATE SEQUENCE bidcreater
    AS BIGINT
    INCREMENT BY 1
    MINVALUE 400001
    NO MAXVALUE;
    ''')
    cur.execute('''CREATE SEQUENCE cidcreater
    AS BIGINT
    INCREMENT BY 1
    MINVALUE 500001
    NO MAXVALUE;
    ''')
    cur.execute('''CREATE SEQUENCE oidcreater
    AS BIGINT
    INCREMENT BY 1
    MINVALUE 600001
    NO MAXVALUE;
    ''')  
    cur.execute('''CREATE SEQUENCE poidcreater
    AS BIGINT
    INCREMENT BY 1
    MINVALUE 700001
    NO MAXVALUE;
    ''')  
def removeData():
    #remove data from all the tables
    cur.execute('delete from publisher')
    cur.execute('delete from genre')
    cur.execute('delete from author')
    cur.execute('delete from book')
    cur.execute('delete from write')
    cur.execute('delete from typeof')
    cur.execute('delete from customer')
    cur.execute('delete from book_order')
    cur.execute('delete from contain')
    cur.execute('delete from utility')
    cur.execute("delete from SHOPPING_CART")
    cur.execute("delete from ORDER_FROM_PUBLISHER")
    cur.execute("delete from PAYMENT_CARD")

def dml():
    # DML
    removeData()
    dropSequence()
    createSequence()

    # publisher table dml
    fake = Faker(['en-CA'])  
    insert_publisher = '''INSERT INTO PUBLISHER VALUES (%s,%s,%s,%s,%s,%s);'''
    publisher_values = []
    for p in publisher_set:
        cur.execute("select nextval('pidcreater')")
        pid = cur.fetchall()[0][0]
        publisher = (p,fake.address(),fakeEmail(p),fake.phone_number(),fakeBank(),pid)
        publisher_values.append(publisher) 
    for p in publisher_values:
        cur.execute(insert_publisher,p) 

    #genre table dml 
    insert_genre = '''INSERT INTO GENRE VALUES (%s, %s);'''
    for g in genre_set:
        cur.execute("select nextval('gidcreater')")
        gid = cur.fetchall()[0][0]
        cur.execute(insert_genre,(g,gid))

    #author table dml 
    insert_author = '''INSERT INTO AUTHOR VALUES (%s,%s)'''
    for a in author_set:
        cur.execute("select nextval('aidcreater')")
        aid = cur.fetchall()[0][0]
        cur.execute(insert_author,(aid,a))

    #book/typeof/write table dml
    for b in bookshelf:
        cur.execute("select nextval('bidcreater')")
        bid = cur.fetchall()[0][0]
        # the percentage need to transfer to publisher when saling
        percentage = randint(50,60)/100
        #get pid
        cur.execute('select pid from publisher where name =%s',(b.publisher,))
        pid=cur.fetchall()[0][0]
        book = (b.title,b.price,b.page_count,b.ISBN,b.storage,bid,pid,percentage)
        print(bid)
        cur.execute('''INSERT INTO BOOK VALUES (%s,%s,%s,%s,%s,%s,%s,%s)''',book)    
        for g in b.genres:
            g = g.strip()
            cur.execute('select gid from genre where genre_type = %s', (g,))
            gid = cur.fetchall()
            cur.execute('INSERT INTO typeof VALUES (%s,%s)',(gid[0][0],bid))    
        for a in b.authors:
            a = a.strip()
            cur.execute('select aid from author where name = %s', (a,))
            aid = cur.fetchall()
            cur.execute('INSERT INTO write VALUES (%s,%s)',(aid[0][0],bid))
    # the last bid inserted
    BID = bid
    conn.commit()

    # add in about 500 fake customer
    names=set()
    for i in range(500):
        names.add(Faker().user_name())
    for n in names:
        cur.execute("select nextval('cidcreater')")
        cid = cur.fetchall()[0][0]
        password= 'password'
        cur.execute("insert into customer values (%s,%s,%s)",(cid,n,password))
        print(cid)
    # the last cid inserted
    CID = cid
    conn.commit()
    # cur.execute('delete from book_order')
    # cur.execute('delete from contain')
    # cur.execute('delete from utility')
    # cur.execute("delete from SHOPPING_CART")
    # cur.execute("delete from ORDER_FROM_PUBLISHER")
    # # add in fake order from Januray to December
    # CID=500496
    # BID=400606
    # cur.execute('DROP SEQUENCE IF EXISTS oidcreater CASCADE;')
    # cur.execute('''CREATE SEQUENCE oidcreater
    # AS BIGINT
    # INCREMENT BY 1
    # MINVALUE 600001
    # NO MAXVALUE;
    # ''')  
    fakeA = Faker(['en-CA'])
    fake = Faker()
    month = 1
    orders=1
    while orders <=3000:
        # refresh the sales amount when enter a new month
        if 3000%250 == 0:
            cur.execute('REFRESH MATERIALIZED VIEW amount_sold_per_month;')
        month = ceil(orders / 250)
        if month in [1,3,5,7,8,10]:
            day=31
        elif month in [4,6,9,11]:
            day=30
        elif month == 2:
            day=28
        else:
            day=9
        print(month)


        # add card
        cvv = randint(100,999)
        cardNum = fakeBank()
        cur.execute("select payment_card.bill_name,payment_card.bill_address from payment_card where card_num=%s",(cardNum,))
        result = cur.fetchall()
        if not result:
            expDate = expDateFaker()   
            billadd = fakeA.address()
            billname = fake.name()
            cardType = 'Visa' if randint(0,1) else 'Master'
            cur.execute('insert into payment_card values (%s,%s,%s,%s,%s,%s)',(cardNum,cvv,expDate,billname,billadd,cardType))
        else:
            billadd = result[0][1]
            billname = result[0][0]

        # add order
        fdate = fake.date_between_dates(date_start=datetime(2022,month,1), date_end=datetime(2022,month,day))    
        fDate = fdate.strftime("%Y-%m-%d")
        carriers = ['Canada Post','UPS','DHL','FedEx','Purolator','eShipper','CanPar']
        l = len(carriers)
        carrier = carriers[randint(0,l-1)]
        a = randint(0,2)
        if a==0: 
            status = 'Your order is in transit'
        elif a==1:
            status = "Your order is being packaged in our warehouse, will be shipped soon"
        else:
            status = 'Your order has been delivered'    
        cur.execute("select nextval('oidcreater')")
        oid = cur.fetchall()[0][0]
        print(oid)
        if randint(0,1):
            shipname = billname
            shipadd = billadd
        else:
            shipname = fake.name()
            shipadd = fakeA.address()

          
        cid = randint(500001,CID)
        # assume there are 1 to 10 books in an order
        cur.execute('insert into book_order values (%s,%s,%s,%s,%s,%s,%s,%s)',(oid,carrier,cid,shipname,shipadd,cardNum,status,fDate))
        # put books in this fake order
        bookcart = set()   
        for i in range (1,3):
            book = randint(400001,BID)         
            bookcart.add(book) 
        for b in bookcart:
            # update the storage
            amount = randint(1,3)
            print(amount)
            cur.execute('select storage from book where bid=%s',(b,))
            storage = cur.fetchall()[0][0]
            if storage < amount:
                continue
            storage -= amount
            print(b)
            print(f"before:{storage}")
            cur.execute('update book set storage = %s where bid=%s',(storage,b)) 
            cur.execute('insert into contain values (%s,%s,%s)',(oid,b,amount))
            
            cur.execute('''SELECT amount_sold_per_month.amount AS amount
                from amount_sold_per_month, book
                where amount_sold_per_month.month = TO_CHAR(date_trunc('month', %s - interval '1' month), 'MM') AND amount_sold_per_month.bid = %s''',(fdate,b))
            last = cur.fetchall()
            if last:
                print(f"last_month :{last[0][0]}")
            else:
                print(last)
            cur.execute('select storage from book where bid=%s',(b,))
            storage = cur.fetchall()[0][0]
            print(f"storage :{storage}")
                    
        orders+=1
    # add fake utility to bookstore db
    month=['01','02','03','04','05','06','07','08','09','10','11','12']
    for m in month:
        #assume the utility per month is 1000.00 to 2000.00
        cost = randrange(100000,200000)/100
        cur.execute('insert into utility values (%s,%s)',(m,cost))  
    conn.commit()

def dropMaterializeView():
    cur.execute('DROP MATERIALIZED VIEW IF EXISTS ct CASCADE;' )
    cur.execute('DROP MATERIALIZED VIEW IF EXISTS book_in_order CASCADE;' )
    cur.execute('DROP MATERIALIZED VIEW IF EXISTS sales_per_author CASCADE;' )
    cur.execute('DROP MATERIALIZED VIEW IF EXISTS sales_per_genre CASCADE;' )
    cur.execute('DROP MATERIALIZED VIEW IF EXISTS sales_to_publisher CASCADE;' )
    cur.execute('DROP MATERIALIZED VIEW IF EXISTS expenditure CASCADE;' )
    cur.execute('DROP MATERIALIZED VIEW IF EXISTS sales_vs_expenditure CASCADE;' )
    cur.execute('DROP MATERIALIZED VIEW IF EXISTS amount_sold_per_month CASCADE;' )
    cur.execute('DROP MATERIALIZED VIEW IF EXISTS book_order_recoerd CASCADE;' )
    
def refreshMaterializeView():
    cur.execute('REFRESH MATERIALIZED VIEW ct;')
    cur.execute('REFRESH MATERIALIZED VIEW book_in_order;')
    cur.execute('REFRESH MATERIALIZED VIEW sales_per_author;')
    cur.execute('REFRESH MATERIALIZED VIEW sales_per_genre;')
    cur.execute('REFRESH MATERIALIZED VIEW sales_to_publisher;')
    cur.execute('REFRESH MATERIALIZED VIEW expenditure;')
    cur.execute('REFRESH MATERIALIZED VIEW sales_vs_expenditure;')
    cur.execute('REFRESH MATERIALIZED VIEW amount_sold_per_month;')
    cur.execute('REFRESH MATERIALIZED VIEW book_order_recoerd;')
    
def dropSequence():
    cur.execute('DROP SEQUENCE IF EXISTS pidcreater CASCADE;')
    cur.execute('DROP SEQUENCE IF EXISTS gidcreater CASCADE;')
    cur.execute('DROP SEQUENCE IF EXISTS aidcreater CASCADE;')
    cur.execute('DROP SEQUENCE IF EXISTS bidcreater CASCADE;')
    cur.execute('DROP SEQUENCE IF EXISTS cidcreater CASCADE;')
    cur.execute('DROP SEQUENCE IF EXISTS oidcreater CASCADE;')
    cur.execute('DROP SEQUENCE IF EXISTS poidcreater CASCADE;')
    

def createMaterializeView():
    # create materialize view
    cur.execute('''CREATE MATERIALIZED VIEW ct 
    AS select book.*,author.aid,author.name as Aname,genre.*,publisher.name as pname from book, write,typeof,genre,author,publisher
                    where book.bid = write.bid and 
                    book.bid = typeof.bid and 
                    typeof.gid = genre.gid and 
                    write.aid = author.aid and
                    book.pid=publisher.pid
    ;''')

    cur.execute('''CREATE MATERIALIZED VIEW book_in_order AS
    select book.title as title, book_order.oid as oid, book_order.carrier as carrier, book_order.SHIPMENT_STATUS as status, contain.amount as amount from BOOK,BOOK_ORDER,CONTAIN
        where book_order.oid = contain.oid and
            contain.bid = book.bid
    ;''')

    # materialized view for sales per author
    cur.execute('''CREATE MATERIALIZED VIEW sales_per_author AS 
    select author.name as name, sum(book.price) as price,TO_CHAR(book_order.sdate, 'MM') as month from contain, book_order, write, book, author
    where book.bid = write.bid and 
    author.aid = write.aid and
    book.bid = contain.bid and
    book_order.oid = contain.oid
    group by TO_CHAR(book_order.sdate, 'MM'),author.name
    order by name desc
    ;''')

    # materialized view for sales per genre
    cur.execute('''CREATE MATERIALIZED VIEW sales_per_genre AS 
    select TO_CHAR(book_order.sdate, 'MM') as month, genre.genre_type as type, sum(book.price) as price from contain, book_order, genre, book, TYPEOF
    where book.bid = typeof.bid and 
    genre.gid = typeof.gid and
    book.bid = contain.bid and
    book_order.oid = contain.oid
    group by TO_CHAR(book_order.sdate, 'MM'), genre.genre_type 
    order by type desc
    ;''')  

    # this is the sales transfered to each publisher by month
    cur.execute('''CREATE MATERIALIZED VIEW sales_to_publisher AS 
    select pname, sum(sales) as sale, TO_CHAR(sdate, 'MM') as month from
    (select publisher.name as pname, publisher.BANK_ACC as pbank, (book.SALESPERCENT_TO_PUB * book.price) as sales, book_order.sdate as sdate
    from contain, book_order, book, publisher
    where book.bid = contain.bid and
    book_order.oid = contain.oid and
    book.pid = publisher.pid)sales_table
    group by pname, TO_CHAR(sdate, 'MM')
    order by pname desc
    ;''')

    # this is the expenditure per month: percentage to publisher + utility
    cur.execute('''CREATE MATERIALIZED VIEW expenditure AS 
    select (utility.cost + pay_pub) as expen, utility.month as month, pay_pub, utility.cost as utility
    from utility, (select sum(sales_to_publisher.sale) as pay_pub, sales_to_publisher.month as pmonth
    from sales_to_publisher
    group by sales_to_publisher.month) pay_pub_table
    where utility.month = pmonth
    order by month asc
    ;''')

    # this is the sales vs expenditure per month:
    cur.execute('''CREATE MATERIALIZED VIEW sales_vs_expenditure AS 
    select expen, utility, pay_pub, sales, (sales - expen) as profit, sale_table.month as month from
    expenditure, 
    (select sum(book.price) as sales, TO_CHAR(book_order.sdate, 'MM') as month from
    book, book_order,contain
    where book.bid= contain.bid and 
    book_order.oid = contain.oid
    group by TO_CHAR(book_order.sdate, 'MM')) sale_table
    where sale_table.month = expenditure.month
    ;''')

    # this is the amount of each book is sold in each month:
    cur.execute('''CREATE MATERIALIZED VIEW amount_sold_per_month AS 
    select bid, sum(sales) as amount, TO_CHAR(sdate, 'MM') as month from
    (select book.bid as bid, 1 as sales, book_order.sdate as sdate                                                                                                                                                                                                                              
    from contain, book_order, book
    where book.bid = contain.bid and
    book_order.oid = contain.oid)sales_table
    group by bid, TO_CHAR(sdate, 'MM')
    order by month asc
    ;''')

    # this shows the details of each order placed with publisher when book storage<10, 
    # including book title ISBN publisher email publisher name
    cur.execute('''CREATE MATERIALIZED VIEW book_order_recoerd AS 
    select publisher.name as pname, publisher.email as pemail, book.title as title, 
    book.ISBN as ISBN, ORDER_FROM_PUBLISHER.order_amount as amount, ORDER_FROM_PUBLISHER.date as date
    from publisher, book, ORDER_FROM_PUBLISHER
    where publisher.pid = ORDER_FROM_PUBLISHER.pid and
    book.bid = ORDER_FROM_PUBLISHER.bid
    ;''')

# order new book from publisher
def trigger_function():
    cur.execute('''CREATE FUNCTION check_book_storage()
        returns TRIGGER
        language plpgsql
        AS
        $$
        DECLARE
            _amount int := (SELECT amount_sold_per_month.amount 
                from amount_sold_per_month,(SELECT BOOK_ORDER.SDATE AS DATE FROM BOOK_ORDER WHERE BOOK_ORDER.OID = NEW.OID) AS SDATE
                where amount_sold_per_month.month = TO_CHAR(date_trunc('month', SDATE.DATE - interval '1' month), 'MM') AND amount_sold_per_month.bid = NEW.bid); 
                
        begin              
            IF (SELECT BOOK.STORAGE
                FROM BOOK
                WHERE BOOK.BID = NEW.BID) < 20
            THEN
                --order from publisher
                WITH SDATE AS
                (SELECT BOOK_ORDER.SDATE AS DATE FROM BOOK_ORDER
                WHERE BOOK_ORDER.OID = NEW.OID)              
                INSERT INTO ORDER_FROM_PUBLISHER 
                SELECT 
                    nextval('poidcreater'),
                    amount_sold_per_month.amount,
                    pid_table.pid,
                    NEW.bid,
                    SDATE.DATE 
                from amount_sold_per_month,(select pid from book where book.bid = new.bid) as pid_table, SDATE
                where amount_sold_per_month.month = TO_CHAR(date_trunc('month', SDATE.DATE - interval '1' month), 'MM') and
                amount_sold_per_month.bid = NEW.bid;
            

                --update the storage of book
                IF EXISTS (SELECT amount_sold_per_month.amount AS amount
                from amount_sold_per_month,(SELECT BOOK_ORDER.SDATE AS DATE FROM BOOK_ORDER WHERE BOOK_ORDER.OID = NEW.OID) AS SDATE
                where amount_sold_per_month.month = TO_CHAR(date_trunc('month', SDATE.DATE - interval '1' month), 'MM') AND amount_sold_per_month.bid = NEW.bid)
                THEN UPDATE BOOK SET STORAGE = storage + _amount WHERE BOOK.BID = NEW.BID;
                ELSE UPDATE BOOK SET STORAGE = storage + 5 WHERE BOOK.BID = NEW.BID;
                END IF;

                  
        END IF;
        RETURN NEW;
        end;
        $$
        ''')
def trigger():
    cur.execute('''CREATE TRIGGER ORDER_PLACE
            AFTER INSERT
            ON CONTAIN
            FOR EACH ROW
            EXECUTE PROCEDURE check_book_storage();    
    ''')

# drop trigger and trigger function
def dropTrigger():
    cur.execute('''DROP TRIGGER IF EXISTS book_update 
        ON BOOK CASCADE;''')
    cur.execute('DROP FUNCTION IF EXISTS check_book_storage() CASCADE;')
    

     
try:
    conn = psycopg2.connect(
        dbname = database,
        user = username,
        password = pwd,
        host = hostname)
    cur=conn.cursor() 

    # uncomment this if want to drop and reintilize this database
    ddl()
    dropMaterializeView()
    createMaterializeView()
    # #drop trigger and recreate trigger
    dropTrigger()
    trigger_function()
    trigger()
    # conn.commit()
    # # uncomment this if want to delete and reinsert intial fake data into database
    dml()   
    # drop materialize view and recreate materialize view
    refreshMaterializeView()
   
    conn.commit()
    print('finished initialization')
except Exception as err:
    print(err)  
finally:
    if cur is not None:
        cur.close()
    if conn is not None:
        conn.close()

            
